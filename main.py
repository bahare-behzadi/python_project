from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.widget import Widget
from kivy.properties import ObjectProperty
from kivy.uix.image import Image
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.uix.image import Image
from kivy.animation import Animation
from kivy.uix.label import Label
from kivy.properties import NumericProperty
import random
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.app import App
from kivy.uix.button import Button
from kivy.lang import Builder
from kivymd.app import MDApp
from kivymd.uix.behaviors import HoverBehavior
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.theming import ThemableBehavior
from kivy.properties import StringProperty, BooleanProperty
from kivy.uix.popup import Popup
from kivy.core.window import Window
import openpyxl
from os import path
import cv2
import pyautogui
from time import sleep
from pygame import *
from pygame.sprite import *
from kivy.core.audio import SoundLoader
from kivy.core.window import Window
from kivy.uix.label import Label
from kivy.graphics import Color, Rectangle


# Define your Screen classes
class FirstWindow(Screen):
    pass

class SecondWindow(Screen):
        
    score = []
    current_question = StringProperty()
    current_choices1 = StringProperty()
    current_choices2 = StringProperty()
    current_choices3 = StringProperty()
    image_source = StringProperty('')
    next_enable = BooleanProperty(False)
    tryagain_enable = BooleanProperty(False) 
    BASE_DIR = path.dirname(path.abspath(__file__))
    file_path = path.join(BASE_DIR, "questions.xlsx")

    wb = openpyxl.load_workbook(file_path)
    sheet_q= wb['Sheet1']
    sheet_ch=wb['Sheet2']

    # Create an empty dictionary and list to store questions
    questions_dict = {}
    added_questions = []
    choice_list=[]
    q_number=[]

# Iterate through rows and add random questions to the dictionary
    while len(questions_dict) < 4:
        row = random.randint(2, sheet_q.max_row) # skip header row
        question = str(sheet_q.cell(row=row, column=1).value)
        answer = str(sheet_q.cell(row=row, column=2).value)
        c1=str(sheet_q.cell(row=row, column=3).value)
        c2=str(sheet_q.cell(row=row, column=4).value)
        c3=str(sheet_q.cell(row=row, column=5).value)
        n=sheet_q.cell(row=row, column=6)
        if question not in added_questions:
            questions_dict[question] = answer
            added_questions.append(question)
            choice_list.append([c1,c2,c3])
    q_number=list(questions_dict.keys())        
    
       
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        
        self.current_question = str(next(iter(self.questions_dict)))
        self.current_choices1 = self.choice_list[0][0]
        self.current_choices2 = self.choice_list[0][1]
        self.current_choices3 = self.choice_list[0][2]
        #self.image = Image(source="Images\question.gif")
        self.box_layout = BoxLayout(orientation='vertical')
        
    
    def on_answer_selected(self, answer):
        girl=''
        colo=[1, 1, 1, 1]
       
            
        if answer[0] == self.questions_dict[self.current_question]:
            result_text = "Correct!"
            girl= "Images\happy.gif"
            colo= color=[0, 1, 0, 1]  
            self.score.append(1) 
              
                   
        else:
            result_text = "Incorrect."
            girl= "Images\cry.gif"
            colo= color=[1, 0, 0, 1] 
            self.score.append(0)
            
            
        img = Image(source=girl)
        label = Label(text=result_text, 
                      font_name="Fonts\ChalktasticItalic-pEwZ.ttf", font_size= 40, color=colo)
        but=Button(text="close", pos_hint= {"center_x": 0.5, "center_y": 0.5}, size_hint=(.5,.4))
        
        
        content = BoxLayout(orientation='vertical')
        content.add_widget(img)
        content.add_widget(label)
        content.add_widget(but)
        popup = Popup(title="Result",
                      content= content,
                      size_hint=(0.5, 0.5), background_color=(0, .7, .3, 1))
        but.bind(on_press=popup.dismiss)
        popup.open()

        self.questions_dict.pop(self.current_question)
        if self.questions_dict:
            
            self.current_question = next(iter(self.questions_dict))
            #self.current_choices = "\n".join(self.questions[self.current_question])
            self.current_choices1 = self.choice_list[self.q_number.index(self.current_question)][0]
            self.current_choices2 = self.choice_list[self.q_number.index(self.current_question)][1]
            self.current_choices3 = self.choice_list[self.q_number.index(self.current_question)][2]

            
            
        else:
            # No more questions
            self.current_question = "your score is: " + str(self.score.count(1))
            self.current_choices1 = "**you finished the game**"
            self.current_choices2 = "**you finished the game**"  
            self.current_choices3 = "**you finished the game**" 
            if  self.score.count(1)>= 3 :
            #off
                self.next_enable=True
                
            else:
            #on
                self.tryagain_enable= True
    def reset_game(self):
        self.questions_dict = {}
        self.added_questions = []
        self.choice_list=[]
        self.q_number=[]
        self.score = []
        self.next_enable = False
        self.tryagain_enable = False
        while len(self.questions_dict) < 4:
            row = random.randint(2, self.sheet_q.max_row) # skip header row
            question =str(self.sheet_q.cell(row=row, column=1).value)
            answer = str(self.sheet_q.cell(row=row, column=2).value)
            c1=str(self.sheet_q.cell(row=row, column=3).value)
            c2=str(self.sheet_q.cell(row=row, column=4).value)
            c3=str(self.sheet_q.cell(row=row, column=5).value)
            n=self.sheet_q.cell(row=row, column=6)
            if question not in self.added_questions:
                self.questions_dict[question] = answer
                self.added_questions.append(question)
                self.choice_list.append([c1,c2,c3])
        self.q_number=list(self.questions_dict.keys()) 
        self.current_question = str(next(iter(self.questions_dict)))
        self.current_choices1 = self.choice_list[0][0]
        self.current_choices2 = self.choice_list[0][1]
        self.current_choices3 = self.choice_list[0][2]

    



    pass 


class ThirdWindow(Screen):
    pass
class FourthWindow(Screen):
    def __init__(self, **kw):
        super().__init__(**kw)
        self.mole_button = None
        self.score = 0
        Window.set_system_cursor('none')
        
        # Create a new image widget for the custom cursor
        self.cursor = Image(source='Images/hunt1.png')
        self.cursor.size_hint = (None, None)
        self.cursor.size = (64, 64)
        self.cursor.pos_hint = {'center_x': 0.5, 'center_y': 0.5}
        self.root = self.cursor
        
        # Bind the image's position to the cursor's position
        Window.bind(mouse_pos=self.on_motion)
        
        self.label_score = Label(text=f"Score: {self.score}", size_hint=(0.3, 0.3), pos_hint={'x': -0.08, 'y': 0.75}, color=(0, 0, 0, 1),font_size="30sp")
        self.add_widget(self.label_score)
        
        Clock.schedule_interval(self.create_mole_button, 1)

    def create_mole_button(self, dt):
        position = [(0.735, .15), (0.74, .5), (0.45, .17), (0.155, .17), (0.44, .50), (0.155, .50), (0.155, .80), (0.44, .8), (.74, .8)]
        random_number = random.randint(0, 8)
        first_element, second_element = position[random_number]
        my_dict = {"Images/r.png": "Images/r2.png", "Images/worm1.png": "Images/worm2.png", "Images/tor1.png": "Images/tor2.png"}
        random_key = random.choice(list(my_dict.keys()))
        if self.mole_button:
            self.remove_widget(self.mole_button)
        self.mole_button = Button(background_normal=random_key, background_down=my_dict[random_key], size_hint=(0.14, 0.14), pos_hint={'x': first_element, 'y': second_element})
        self.mole_button.bind(on_release=self.on_mole_button_release)
        self.add_widget(self.mole_button)

    def on_mole_button_release(self, button):
        self.score += 1
        self.label_score.text = f"Score: {self.score}"
        self.remove_widget(button)

    def on_motion(self, window, pos):
        self.cursor.pos = (pos[0] - self.cursor.width / 2, pos[1] - self.cursor.height / 2)

        

    
    
    

class WindowManager(ScreenManager):
    pass
  



    def dismiss_popup(self):
        self.dismiss() 
   
    pass 

            
    pass
class Background(Widget):
    cloud_texture= ObjectProperty(None)
    floor_texture= ObjectProperty(None)
    bird_texture= ObjectProperty(None)
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # creating texture (moving clouds)
        self.cloud_texture = Image(source='Images\clouds4.png').texture
        self.cloud_texture.wrap='repeat'
        self.cloud_texture.uvsize = (Window.width/self.cloud_texture.width, -1)
        self.floor_texture = Image(source='Images\grass.png').texture
        self.floor_texture.wrap='repeat'
        self.floor_texture.uvsize=(Window.width/self.floor_texture.width, -1)
        
    def scroll_textures(self, time_passed):
       self.cloud_texture.uvpos = ((self.cloud_texture.uvpos[0]+ time_passed)%Window.width,self.cloud_texture.uvpos[1])
       texture= self.property('cloud_texture')
       texture.dispatch(self)
       self.floor_texture.uvpos = ((self.floor_texture.uvpos[0]+ time_passed/3)%Window.width,self.floor_texture.uvpos[1])
       texture= self.property('floor_texture')
       texture.dispatch(self)
class MyButton(Button):
    def on_enter(self):
        Window.set_system_cursor("hand")

    def on_leave(self):
        Window.set_system_cursor("arrow")
    

    

class MyImage(Image):
    pass       
    def on_touch_down(self, touch):
         # Create an animation to move the image to the right over 5 seconds
        anim = Animation(pos_hint={'right': 1, 'center_y': 0.8}, duration=2)
        anim+=Animation(opacity=0, duration=7.2)
        anim.start(self)  # Start the animation on the image
class MyLabel(Label):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.contex=["hi!","this","is","a","test"]
        self.index = 0  # Keep track of the current index
        self.update_text()  # Call the update_text function once to start the loop

    def update_text(self, *args):
        # Check if we reached the end of the list
        if self.index >= len(self.contex):
            return
        
        # Update the label text with the current item of the list
        self.text = self.contex[self.index]

        # Schedule a callback to update the label text after a delay of 2 seconds
        Clock.schedule_once(self.update_text, 2)

        # Increment the index to move to the next item in the list
        self.index += 1

    


   

   



# Define your App class
class MyApp(MDApp):
    

    def build(self):
        # Create a ScreenManager and add each screen to it
        sm = WindowManager()
        sm.add_widget(FirstWindow(name='first'))
        sm.add_widget(SecondWindow(name='second'))
        sm.add_widget(ThirdWindow(name='third'))
        sm.add_widget(FourthWindow(name='fourth'))
        
        return sm
    def on_start(self):
        Clock.schedule_interval(self.root.get_screen('first').ids.B_L.scroll_textures, 1/3.)
    pass
MyApp().run()